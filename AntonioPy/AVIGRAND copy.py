# python ou py -m venv env
# .\env\Scripts\activate
# deactivate
# pip install os
# pip install pyxlsb
# pip install tabula-py
# pip install pandas
# pip install numpy
# pip install PyPDF2
# pip install tqdm
# pip uninstall tabula-py
# pip uninstall tabula
# sudo pip uninstall tabula-py
# sudo pip uninstall tabula
#from os import *
#from tabula import *
#from tabula import read_pdf_with_template
#from openpyxl import load_workbook
#import os
#import pandas as pd
#import numpy as np
#import PyPDF2 as pypdf
#import time
#from tqdm import tqdm
import os
from tabula import *
from tabula import read_pdf_with_template
import pandas as pd
import numpy as np
import PyPDF2 as pypdf
import time
from tqdm import tqdm
from openpyxl import load_workbook


t1 = time.time()
cols = []
tabelas = {}
files = ["PDF_AVIGRAND_T", "ARQUIVOS_EXCEL", "MODELOS_PDF"]

chaves = [
    "INFORMAÇÕES GERAIS",
    "CONDENAÇÕES DE CAUSAS AGROPECUÁRIA",
    "PESO SEMANAL",
    "ORGEM DO LOTE",
    "ABATES",
    "MOVIMENTAÇÃO DE RAÇÕES",
    "HISTÓRICO DOS PEDIDOS ANTERIORES",
]

for dir_ in files:
    if not os.path.exists(dir_):
        os.system(f"mkdir {dir_}")
        # print(system)

def reader_pages(Nome_pdf):
    for arquivo_pdf in os.listdir(files[0]):
        pdf = pypdf.PdfReader(f"{files[0]}/{Nome_pdf}")
    return int(len(pdf.pages))

def remover_dois_pontos(frase):
    frase_sem_dois_pontos = frase.replace(':', '')
    return frase_sem_dois_pontos

def manipule_array(arr):
    global colum
    char_l = []

    chave_atual = None
    qtd_it_arr = len(arr)

    for col in range(qtd_it_arr):

        df = pd.DataFrame(arr[col])
        df.insert(0, f"CHAVE", f"{name_file}")
        lines, coluns = df.shape
        colum = df.columns[1]
        colum = str(colum).replace(":", "")

        if colum in chaves:
            tabelas[colum] = []
            colum = colum
            cols.append(colum)
        for line in range(0, lines):

            line_t = list(df.iloc[line].values)
            char_l = str(line_t[1]).replace(":", "")



            if char_l in chaves:
                chave_atual = char_l
                tabelas[char_l] = []
                cols.append(chave_atual)
            cols.append(line_t)

    df = pd.DataFrame(cols)

    # Verifique se o arquivo existe
    if os.path.exists("TABELA_1.xlsx"):
        book = load_workbook("TABELA_1.xlsx")
        writer = pd.ExcelWriter("TABELA_1.xlsx", engine='openpyxl') 
        writer.book = book
    else:
        writer = pd.ExcelWriter("TABELA_1.xlsx", engine='openpyxl')

    df.to_excel(writer, index=False)
    writer.save()
    writer.close()

    return


def main():
    global name_file

    # Loop para converter cada PDF em um DataFrame
    try:
        for arquivo in tqdm(listdir(files[0]), desc="PROCESSANDO ARQUIVOS"):

            if arquivo.endswith(".pdf"):
                pdf_num_pages = reader_pages(arquivo)
                nome_arquivo = path.splitext(arquivo)[0]
                name_file = nome_arquivo

                if pdf_num_pages == 3:
                    templateJson = "T3.tabula-template.json"
                if pdf_num_pages == 4:
                    templateJson = "T4.tabula-template.json"
                if pdf_num_pages == 5:
                    templateJson = "T5.tabula-template.json"

                dfs = read_pdf_with_template(
                    path.join(files[0], arquivo), f"{files[2]}/{templateJson}", stream=True)
                manipule_array(dfs)
               # print(dfs)
    except Exception as err:
        print("Erro :",  err)


main()
t2 = time.time() - t1
print("Tempo de exec: ", t2)
# print(reader_pages())
